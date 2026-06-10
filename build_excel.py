"""
Tutor Schedule Diff — build report
Usage: called with wb1 (דרוש תיקון), wb2 (ייצוא שיטס) already loaded and fixed.
All config is discovered dynamically from the files themselves.
"""

import difflib, re, math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, time as dt_time
from collections import defaultdict, Counter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — edit these when new branches/tutors/subjects are added
# ═══════════════════════════════════════════════════════════════════════════════

# Known branch locations — extend when new branches appear
KNOWN_LOCATIONS = {
    'אלתרמן','תל מונד','גאולים','ניצני עוז','עמי אסף','עמיאסף',
    'ירקון','צורן','זום','צור יצחק','רופין','חגור','מתן','כפר יונה','שרונה','ניצני',
}

# Aliases → canonical name
LOCATION_NORMALIZE = {
    'שרונה':   'כפר יונה',
    'עמיאסף':  'עמי אסף',
    'ניצני':   'ניצני עוז',
}

# Colors per branch (from the example file col B)
BRANCH_COLORS = {
    'אלתרמן':'CCC8D3','תל מונד':'ACA9B5','גאולים':'8F8E99',
    'ניצני עוז':'919398','עמי אסף':'93A0A5','ירקון':'A5B2BA',
    'צורן':'B7C4CA','זום':'C3D6D3','צור יצחק':'D7E6DA',
    'רופין':'C5D1BD','חגור':'AEB9A3','מתן':'B9B79C',
    'כפר יונה':'CCCBAF','default':'F0EFEF',
}
BRANCHES_ORDER = [
    'אלתרמן','תל מונד','גאולים','ניצני עוז','עמי אסף',
    'ירקון','צורן','זום','צור יצחק','רופין','חגור','מתן','כפר יונה',
]

# Subject keywords in time-slot labels (used for per-slot subject detection)
SUBJECT_KEYWORDS = {
    'מתמטיקה': ['מתמטיקה', 'מתמטיק'],
    'אנגלית':  ['אנגלית', 'אנגל'],
}

# ── Student name cleaning ──────────────────────────────────────────────────────
# Hebrew grade letters — when found after a name, mark end of name
HEBREW_GRADES = {'א','ב','ג','ד','ה','ו','ז','ח','ט','י','יא','יב','יג'}
# Words that are definitively NOT part of a student name
NOISE_WORDS = {
    'שדי','חמד','ויתרתי','עד','בגלל','בדיקה','רפואית','לה','על','חיסור',
    'בית','יצחק','שרונה','גאולים','ירקון','צורן','זום','רופין','חגור','מתן',
    'תל','מונד','אלתרמן','ניצני','עוז','עמי','אסף',
}

# Fallback tutor→subject map (used ONLY when no label exists in the slot)
# A tutor in BOTH sets will follow slot labels; if no label → 'מתמטיקה'
MATH_TUTORS = {
    'אדיר','אלעד','דעאל','יניב','מיכאל','מיקה','נדב','ניצן',
    'עמית','עמרי','קרן','אלכס','דקל','גל',
}  # נועה/נעה teach English only — removed from math
ENG_TUTORS = {
    'גריפין','נגה','סיגל','סמרה','ענבר','קמיל','נועה','ניצן','דקל','גל','נעה',
}
DEFAULT_SUBJECT = 'מתמטיקה'   # fallback when tutor not in either set

# Font for tutors row (must match the template file)
MAIN_FONT = 'Calibri'
HEADER_COLOR = '1F2D3D'

# ═══════════════════════════════════════════════════════════════════════════════
# LOCATION EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def find_known_location(text):
    """Find the FIRST known location token in a text fragment.
    Returns canonical name or None. Always checks multi-word before single-word."""
    text = text.strip()
    for loc in sorted([l for l in KNOWN_LOCATIONS if ' ' in l], key=len, reverse=True):
        if loc in text:
            return LOCATION_NORMALIZE.get(loc, loc)
    for word in re.split(r'[\s\-/\+\?\!]+', text):
        w = word.strip()
        if w in KNOWN_LOCATIONS:
            return LOCATION_NORMALIZE.get(w, w)
    return None

def extract_location(raw):
    """Extract branch/location from a raw tutor+location string.
    Handles all known formats robustly. Returns None if no known location found.
    To extend: add new entries to KNOWN_LOCATIONS / LOCATION_NORMALIZE above."""
    raw = str(raw).strip()
    # Remove "X מחליפ(ה) את Y" prefix
    raw = re.sub(r'מחליפ[הת]?\s+את\s+[\u05d0-\u05ea\s]+', '', raw)
    # Remove parenthetical notes like "(במקום עמית)", "(שיעורי השלמה)"
    raw = re.sub(r'\([^)]*\)', '', raw).strip()

    # Strategy 1: split on newline, process first segment's dash parts
    segments = re.split(r'\n', raw)
    dash_parts = re.split(r'\s*-\s*|\s+-|-\s+', segments[0])
    if len(dash_parts) > 1:
        for part in dash_parts[1:]:
            loc = find_known_location(part)
            if loc: return loc

    # Strategy 2: remaining newline segments often contain location
    for seg in segments[1:]:
        loc = find_known_location(seg)
        if loc: return loc

    # Strategy 3: no dash — words after tutor first name
    words = raw.split()
    if len(words) >= 2:
        rest = [w for w in words[1:] if not (w.startswith('ו') and len(w) > 1 and w not in KNOWN_LOCATIONS)]
        loc = find_known_location(' '.join(rest))
        if loc: return loc

    return None

# ═══════════════════════════════════════════════════════════════════════════════
# PARSING UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def norm(n):
    if not n: return ''
    return re.sub(r'\s+', ' ', re.sub(r"['\u05f3\u2019\u05f4\"]", '', str(n))).strip()

def clean_student_name(raw):
    """Extract core student name, stripping grades, dates, notes, location suffixes.
    Examples:
      'עומר פילוס עד 10.5 (לבנתיים)' → 'עומר פילוס'
      'מורן אשר 21.1 ח הדר עם'        → 'מורן אשר'
      'שליו פרידלנדר-ח'               → 'שליו פרידלנדר'
    To extend: add noise words to NOISE_WORDS above."""
    if not raw: return ''
    s = str(raw).strip()
    s = s.replace('?', '').strip()
    s = re.sub(r'\([^)]*\)', '', s).strip()           # remove (notes)
    s = re.sub(r'[״׳]', '', s).strip()                  # remove Hebrew geresh
    s = re.sub(r'-([א-ת]{1,2})(?=\s|$)', r' \1', s)  # "פרידלנדר-ח" → "פרידלנדר ח"
    words = s.split()
    name_words = []
    for word in words:
        w = word.strip("'-,.")
        if not w: continue
        if w in HEBREW_GRADES: break
        if re.search(r'\d', w): break
        if w in NOISE_WORDS: break
        if len(w) > 12 and len(name_words) >= 2: break
        name_words.append(word)
    result = ' '.join(name_words).strip(" '-,.")
    return re.sub(r'\s+', ' ', result)


def is_time_marker(val):
    """Detect time-marker cells generically.
    Strategy: any cell containing HH:MM is treated as a time marker,
    provided it's short enough (≤25 chars) to not be a student name + note.
    Real student names never contain HH:MM patterns.
    Handles semicolons as time separators (e.g. '15;00' typed instead of '15:00').

    Examples that match:
      14:30:00, 14:30 - 16:00, 14:30 אנגלית, 13:30- רופין,
      צורן- 17:00, 13:30 - שיעור השלמה, 15:00 קבוצה ב', 14:00 תלמידי ה-ו, 15;00

    Examples that don't match:
      יהל בכורי, עמרי לבקוביץ (no HH:MM)"""
    if isinstance(val, dt_time): return True
    v = str(val).strip().replace(';', ':')  # normalize semicolons (common typo)
    # Pure time formats
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?(\s*[-–]\s*\d{1,2}:\d{2})?$', v): return True
    # Any cell with HH:MM + short annotation (subject, location, note, anything)
    if re.search(r'\d{1,2}:\d{2}', v) and len(v) <= 25:
        return True
    return False

def xtime(val):
    if isinstance(val, dt_time): return f"{val.hour:02d}:{val.minute:02d}"
    m = re.search(r'(\d{1,2})[;:](\d{2})', str(val))
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else ''

def detect_subject_label(val):
    """Return subject name if val contains a subject keyword, else None."""
    v = str(val)
    for subj, keywords in SUBJECT_KEYWORDS.items():
        if any(k in v for k in keywords): return subj
    return None

def parse_date_flexible(s):
    """Parse date from many formats. Year-agnostic: infers from context or defaults to current year."""
    s = str(s).strip()
    # Full date: 10/05/2026 or 10.05.2026
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m: return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # Short date: 10.05 — infer year from context (set at parse time)
    m = re.search(r'(\d{1,2})\.(\d{1,2})', s)
    if m:
        day, month = int(m.group(1)), int(m.group(2))
        # Use the inferred year (set by caller)
        return datetime(_inferred_year, month, day)
    return None

_inferred_year = datetime.now().year  # updated dynamically per file

def infer_year_from_workbook(wb):
    """Find the most common year in the workbook's dates to avoid hardcoding 2026."""
    years = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for val in row:
                if val:
                    m = re.search(r'[./](\d{4})', str(val))
                    if m: years.append(int(m.group(1)))
    if years:
        return Counter(years).most_common(1)[0][0]
    return datetime.now().year

def parse_date(s, year=None):
    s = str(s).strip()
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', s)
    if m: return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.search(r'(\d{1,2})\.(\d{1,2})', s)
    if m:
        yr = year or _inferred_year
        return datetime(yr, int(m.group(2)), int(m.group(1)))
    return None

def names_match(a, b):
    """Fuzzy name matching. Uses SequenceMatcher — never character-set overlap."""
    def clean(n):
        n = re.sub(r"['\u05f3\u2019\u05f4\"\-\(\)\[\]]", '', str(n))
        n = re.sub(r'\s+[חטיכזאבדה]\d*$', '', n.strip())
        return re.sub(r'\s+', ' ', n).strip()
    ac, bc = clean(a), clean(b)
    if not ac or not bc: return False
    if ac == bc: return True
    an, bn = ac.replace(' ', ''), bc.replace(' ', '')
    if an == bn: return True
    aw, bw = ac.split(), bc.split()
    # Same first name → lower threshold (handles last-name typos)
    if aw and bw and aw[0] == bw[0]:
        return difflib.SequenceMatcher(None, an, bn).ratio() >= 0.70
    return difflib.SequenceMatcher(None, an, bn).ratio() >= 0.85

def tutor_canon(raw):
    """Extract canonical first name from raw tutor+location string."""
    m = re.search(r'מחליפ[הת]?\s+את\s+([\u05d0-\u05ea]+)', str(raw))
    if m: return m.group(1)
    first = re.split(r'[-–\n]', str(raw))[0].strip()
    first = re.sub(r'\s+ו[\u05d0-\u05ea]+.*', '', first).strip()
    return first.split()[0] if first.split() else first

_NOT_IN_SYS_RE = re.compile(
    r'לא\s+(?:נכנס[ה]?|מתעדכן[ת]?|הצלחתי\s+להכניס)',
    re.UNICODE,
)

def parse_slots(ws, col, start_row, first_time='', first_subject=None, strip_sys_note=False):
    """Parse a column into time slots. Returns {time: {'students': set, 'subject': str|None}}.
    strip_sys_note=True: used for yitzua — strip 'לא נכנס למערכת' notes and treat as regular students."""
    current_t = first_time
    current_s = first_subject
    slots = {}
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row, col)
        if cell.value is None: continue
        val = cell.value
        vs = str(val).strip()
        if not vs: continue
        if is_time_marker(val):
            t = xtime(val)
            if t:
                current_t = t
                lbl = detect_subject_label(vs)
                if lbl: current_s = lbl  # subject propagates to subsequent rows
        else:
            if len(vs) > 50: continue
            name = clean_student_name(vs)
            # "לא נכנס/ה למערכת" / "לא הצלחתי להכניס" etc.
            if _NOT_IN_SYS_RE.search(vs):
                extracted = clean_student_name(_NOT_IN_SYS_RE.split(vs)[0])
                if extracted and current_t is not None:
                    if current_t not in slots:
                        slots[current_t] = {'students': {}, 'subject': current_s}
                    if strip_sys_note:
                        # yitzua: ignore the note, treat as a regular student so the
                        # normal comparison can report them as 'להוסיף' if missing from darush
                        slots[current_t]['students'][extracted] = get_attendance(
                            cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else '00000000'
                        )
                    else:
                        # darush: flag explicitly so the diff shows 'להזין למערכת'
                        slots[current_t]['students'][extracted] = 'לא מתעדכן'
                continue
            # Skip: strikethrough text — student notified absence in advance
            if cell.font and getattr(cell.font, 'strike', False):
                continue
            if not name:  # empty after cleaning
                continue
            if current_t not in slots:
                slots[current_t] = {'students': {}, 'subject': current_s}
            slots[current_t]['students'][name] = get_attendance(
                cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else '00000000'
            )
    return slots

def is_junk(name):
    """Filter out non-student values that slipped into student rows."""
    if not name: return True
    if re.match(r'^=', name): return True                               # Excel formula
    if re.match(r'^\d{1,2}:\d{2}', name): return True                  # time value
    if re.search(r'\d{1,2}:\d{2}', name) and len(name) < 15: return True
    # Location names that appear without context
    locs = set(KNOWN_LOCATIONS) | set(LOCATION_NORMALIZE.keys())
    return name.strip() in locs

# ── Attendance status ──────────────────────────────────────────────────────────
GREEN_RGBS = {'FF00FF00','FFB6D7A8','FFD9EAD3','FF9AC47A','FF64FF00'}
RED_RGBS   = {'FFFF0000','FFEE0000','FFEA9999','FFF4CCCC'}

def get_attendance(rgb):
    """Convert cell RGB → attendance status.
    Green  → הגיע | Red → לא הגיע | Anything else → לא סומן (needs marking)"""
    r = str(rgb).upper() if rgb else '00000000'
    if r in GREEN_RGBS: return 'הגיע'
    if r in RED_RGBS:   return 'לא הגיע'
    return 'לא סומן'   # white, no color, yellow, or any unknown color

def tutor_subject(tutor_first, slot_lbl):
    """Determine subject. Slot label always wins over fallback map."""
    if slot_lbl: return slot_lbl
    if tutor_first in ENG_TUTORS and tutor_first not in MATH_TUTORS: return 'אנגלית'
    if tutor_first in MATH_TUTORS and tutor_first not in ENG_TUTORS: return 'מתמטיקה'
    return DEFAULT_SUBJECT

def _time_to_min(t):
    """Convert 'HH:MM' string to total minutes. Returns -9999 on error."""
    try: h, m = str(t).split(':')[:2]; return int(h) * 60 + int(m)
    except: return -9999

# ═══════════════════════════════════════════════════════════════════════════════
# ROW HEIGHT CALCULATION
# ═══════════════════════════════════════════════════════════════════════════════

def note_row_height(text, chars_per_line=58, line_pt=14, pad=8):
    lines = max(1, math.ceil(len(str(text)) / chars_per_line))
    return lines * line_pt + pad

def tutors_row_height(text, chars_per_line=40, line_pt=50, pad=12):
    lines = max(1, math.ceil(len(str(text)) / chars_per_line))
    return lines * line_pt + pad

def _fit_font(text, col_units, max_lines, char_ratio=0.7, px_per_unit=7.2):
    """Largest font size (pt) where text fits within max_lines in the column.
    Uses char_ratio=0.7 calibrated for Calibri Bold Hebrew."""
    col_px = col_units * px_per_unit
    for size in range(22, 7, -1):
        chars_per_line = max(1, int(col_px / (size * char_ratio)))
        if math.ceil(len(text) / chars_per_line) <= max_lines:
            return size
    return 8

def _cell_height(text, col_units, font_size, char_ratio=0.7, px_per_unit=7.2):
    col_px = col_units * px_per_unit
    chars_per_line = max(1, int(col_px / (font_size * char_ratio)))
    lines = math.ceil(len(text) / chars_per_line)
    return int(lines * font_size * 1.6 + 10)

def calc_header_font_size(subject_text, tutors_text):
    """Return (subject_font_size, tutors_font_size, row_height_pt).
    Subject cell (col B, 14 units) must fit in exactly 1 line.
    Tutors cell (col D:G, 96 units) can wrap to max 3 lines."""
    sz_b  = _fit_font(subject_text, col_units=14, max_lines=1)
    sz_dg = _fit_font(tutors_text,  col_units=96, max_lines=3)
    h_b   = _cell_height(subject_text, 14, sz_b)
    h_dg  = _cell_height(tutors_text,  96, sz_dg)
    row_h = max(h_b, h_dg) + 8
    return sz_b, sz_dg, int(row_h)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN COMPARISON FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════

def run_comparison(wb1, wb2):
    """Run full comparison. Returns (differences, tutor_full_map, date_range_str)."""
    global _inferred_year

    # Infer year dynamically from the files
    _inferred_year = infer_year_from_workbook(wb1) or infer_year_from_workbook(wb2) or datetime.now().year
    print(f"Inferred year: {_inferred_year}")

    ws1 = wb1.active

    # ── Parse דרוש תיקון ──
    date_cols = {}
    for col in range(1, ws1.max_column + 1):
        v = ws1.cell(1, col).value
        if v: date_cols[col] = v

    darush = {}
    tutor_full_map = {}
    all_darush_dates = []
    date_list = sorted(date_cols.keys())

    for i, sc in enumerate(date_list):
        ec = date_list[i+1] if i+1 < len(date_list) else ws1.max_column + 1
        date_obj = parse_date(str(date_cols[sc]), _inferred_year)
        if not date_obj: continue
        all_darush_dates.append(date_obj)
        for col in range(sc, ec):
            tv = ws1.cell(2, col).value
            if not tv: continue
            tf = norm(tv)
            t1 = tf.split()[0]
            tutor_full_map[t1] = tf
            ft = xtime(ws1.cell(4, col).value or '')
            for t, sd in parse_slots(ws1, col, 5, ft).items():
                key = (date_obj, t1, t)
                if key not in darush: darush[key] = {}
                darush[key].update(sd['students'])  # {name: attendance}

    darush_dates = set(k[0] for k in darush)
    darush_tutor_dates = set((k[0], k[1]) for k in darush)
    darush_session_count = Counter((k[0], k[1]) for k in darush)

    # ── Parse ייצוא שיטס ──
    yitzua = []
    empty_sessions = []
    yitzua_sheet_dates = set()   # all dates covered by any yitzua sheet header
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        day_cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v and 'יום' in str(v): day_cols[c] = str(v)
        for v in day_cols.values():
            d = parse_date(v, _inferred_year)
            if d: yitzua_sheet_dates.add(d)
        for col in range(1, ws.max_column + 1):
            raw = str(ws.cell(2, col).value or '').strip()
            if not raw: continue
            canon = tutor_canon(raw)
            loc = extract_location(raw)
            dl = next((v for c, v in sorted(day_cols.items(), reverse=True) if col >= c), '')
            date_obj = parse_date(dl, _inferred_year)
            if not date_obj or date_obj not in darush_dates: continue
            r3 = ws.cell(3, col).value
            if r3 and isinstance(r3, str) and re.search(r'[\u05d0-\u05ea]{2,}', r3) and not re.search(r'\d+:\d+', r3):
                ft = ''; fs = detect_subject_label(r3)
            else:
                ft = xtime(r3) if r3 else ''
                fs = detect_subject_label(str(r3)) if r3 else None
            col_had_students = False
            for t, sd in parse_slots(ws, col, 4, ft, fs, strip_sys_note=True).items():
                if sd['students']:
                    col_had_students = True
                    subj = tutor_subject(canon, sd['subject'])
                    full = tutor_full_map.get(canon, canon)
                    yitzua.append({'date': date_obj, 'canon': canon, 'tutor': full,
                                   'time': t, 'students': sd['students'],
                                   'subject': subj, 'location': loc})
            if not col_had_students and (date_obj, canon) in darush_tutor_dates:
                full = tutor_full_map.get(canon, canon)
                empty_sessions.append({'date': date_obj, 'canon': canon, 'tutor': full,
                                       'subject': tutor_subject(canon, None), 'location': loc,
                                       'count': darush_session_count[(date_obj, canon)]})

    # ── Compare ──
    differences = []
    missing_darush_sessions = []
    processed_darush_keys = set()
    for sy in yitzua:
        key = (sy['date'], sy['canon'], sy['time'])
        if key not in darush:
            # Try exact time + fuzzy tutor name
            alts = {k: v for k, v in darush.items()
                    if k[0] == sy['date'] and k[2] == sy['time'] and
                    (sy['canon'] in k[1] or k[1] in sy['canon'])}
            if not alts:
                # Try fuzzy time (±30 min) + exact canon; pick closest match
                sy_min = _time_to_min(sy['time'])
                alts = {k: v for k, v in darush.items()
                        if k[0] == sy['date'] and k[1] == sy['canon'] and
                        sy_min >= 0 and 0 < abs(_time_to_min(k[2]) - sy_min) <= 30}
            if not alts:
                # Session exists in yitzua but time slot is completely absent from darush
                if (sy['date'], sy['canon']) in darush_tutor_dates:
                    missing_darush_sessions.append({
                        'date': sy['date'], 'canon': sy['canon'],
                        'tutor': sy['tutor'], 'time': sy['time'],
                        'subject': sy['subject'], 'location': sy['location'],
                    })
                continue
            key = min(alts, key=lambda k: abs(_time_to_min(k[2]) - _time_to_min(sy['time'])))
        processed_darush_keys.add(key)
        d_sts = darush[key]
        y_sts = sy['students']
        md, my = set(), set()
        for yn in y_sts:
            for dn in d_sts:
                if dn in md: continue
                if names_match(yn, dn):
                    md.add(dn); my.add(yn)
                    # Student absent in yitzua but marked present in darush → needs correction
                    if y_sts[yn] == 'לא הגיע' and d_sts.get(dn) == 'הגיע':
                        differences.append({'tutor': sy['tutor'], 'date': sy['date'], 'time': sy['time'],
                                             'type': 'לסמן באדום', 'student': dn,
                                             'attendance': None,
                                             'subject': sy['subject'], 'location': sy['location']})
                    break
        for yn in y_sts:
            if yn not in my and not is_junk(yn):
                differences.append({'tutor': sy['tutor'], 'date': sy['date'], 'time': sy['time'],
                                     'type': 'להוסיף', 'student': yn,
                                     'attendance': y_sts[yn],
                                     'subject': sy['subject'], 'location': sy['location']})
        for dn in d_sts:
            if dn not in md and not is_junk(dn):
                gap_type = 'לא מתעדכן' if d_sts[dn] == 'לא מתעדכן' else 'להסיר'
                differences.append({'tutor': sy['tutor'], 'date': sy['date'], 'time': sy['time'],
                                     'type': gap_type, 'student': dn,
                                     'attendance': None,
                                     'subject': sy['subject'], 'location': sy['location']})

    # ── Sessions in darush but absent from yitzua → all students are "להסיר" ──
    # Skip (date, canon) pairs handled as empty yitzua columns, and dates with no yitzua sheet
    empty_covered = set((es['date'], es['canon']) for es in empty_sessions)
    for key in darush:
        if key in processed_darush_keys:
            continue
        date_obj, canon, t = key
        if date_obj not in yitzua_sheet_dates:
            continue   # yitzua file has no sheet for this week — not comparable
        if (date_obj, canon) in empty_covered:
            continue
        d_sts = darush[key]
        full = tutor_full_map.get(canon, canon)
        subj = tutor_subject(canon, None)
        for dn in d_sts:
            if not is_junk(dn):
                gap_type = 'לא מתעדכן' if d_sts[dn] == 'לא מתעדכן' else 'להסיר'
                differences.append({'tutor': full, 'date': date_obj, 'time': t,
                                     'type': gap_type, 'student': dn,
                                     'attendance': None,
                                     'subject': subj, 'location': None})

    # ── Date range ──
    all_dates = [d['date'] for d in differences if d['date']]
    if not all_dates: all_dates = all_darush_dates
    date_from = min(all_dates).strftime('%d.%m.%Y')
    date_to   = max(all_dates).strftime('%d.%m.%Y')
    date_range_str = f"{date_from} — {date_to}"

    return differences, tutor_full_map, date_range_str, empty_sessions, missing_darush_sessions

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_report(differences, tutor_full_map, date_range_str, output_path, empty_sessions=None, missing_darush_sessions=None):
    def sfill(h): return PatternFill(fill_type='solid', fgColor=h)
    def fnt(bold=False, color='222222', size=10, italic=False):
        return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
    thin = Side(style='thin', color='C0C0C0')
    THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
    def align(h='right', wrap=True):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap, readingOrder=2)

    ATTENDANCE_ICON = {'הגיע': '🟢', 'לא הגיע': '🔴', 'לא סומן': '⬜', None: ''}

    def format_student(d):
        """Format student name with attendance icon."""
        icon = ATTENDANCE_ICON.get(d.get('attendance'), '')
        name = d['student']
        att  = d.get('attendance')
        if att == 'לא סומן':
            return f"{name} {icon}יש לסמן"
        return f"{name} {icon}" if icon else name

    def build_note(ds):
        adds     = [d for d in ds if d['type'] == 'להוסיף']
        rems     = [d for d in ds if d['type'] == 'להסיר']
        marks    = [d for d in ds if d['type'] == 'לסמן באדום']
        not_sys  = [d for d in ds if d['type'] == 'לא מתעדכן']
        parts = []
        if adds:
            students_str = ', '.join(format_student(d) for d in adds)
            parts.append('להוסיף את: ' + students_str)
        if rems:
            students_str = ', '.join(d['student'] for d in rems)
            parts.append('להסיר את: ' + students_str)
        if marks:
            students_str = ', '.join('🔴 ' + d['student'] for d in marks)
            parts.append('יש לסמן באדום: ' + students_str)
        if not_sys:
            students_str = ', '.join('⚠️ ' + d['student'] for d in not_sys)
            parts.append('להזין למערכת: ' + students_str)
        return ' | '.join(parts)

    def branch_fill(loc):
        h = 'FF' + BRANCH_COLORS.get(loc, BRANCH_COLORS['default'])
        return sfill(h)

    WARN_FILL = sfill('FFFFD54F')
    MISS_FILL = sfill('FFEF9A9A')

    all_tutors_full = sorted(set(tutor_full_map.values()))
    math_tutors = [t for t in all_tutors_full if t.split()[0] in MATH_TUTORS]
    eng_tutors  = [t for t in all_tutors_full if t.split()[0] in ENG_TUTORS]
    other       = [t for t in all_tutors_full if t not in math_tutors and t not in eng_tutors]
    math_tutors += other

    def make_sheet(wb, title, sheet_diffs, tutors, empty_sessions_map=None, missing_ds_map=None):
        ws = wb.create_sheet(title=title)
        ws.sheet_view.rightToLeft = True
        for col, w in zip(range(1, 8), [3, 14, 3, 16, 56, 10, 14]):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Row 1: date range title
        ws.row_dimensions[1].height = 32
        c = ws.cell(1, 2, f'דוח פערים — {title} | {date_range_str}')
        c.font = Font(name=MAIN_FONT, bold=True, size=20, color='FFFFFF')
        c.fill = sfill('FF1A3C5E')
        c.alignment = align('right', False)
        ws.merge_cells('B1:G1')

        # Row 2: subject + tutors (felix007, auto-size to prevent overflow, wrap)
        shorts = sorted(t.split()[0] for t in tutors)
        tutors_text = 'מדריכים שנבדקו: ' + ', '.join(shorts) + '.'
        sz_b, sz_dg, row2_h = calc_header_font_size(title, tutors_text)
        ws.row_dimensions[2].height = row2_h

        c = ws.cell(2, 2, title)
        c.font = Font(name=MAIN_FONT, bold=True, size=sz_b, color=HEADER_COLOR)
        c.fill = sfill('FFF5F5F5')
        # shrink_to_fit forces Excel to auto-reduce font size until text fits in B2 — no overflow ever
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=False, shrink_to_fit=True, readingOrder=2)

        c = ws.cell(2, 4, tutors_text)
        c.font = Font(name=MAIN_FONT, bold=True, size=sz_dg, color=HEADER_COLOR)
        c.fill = sfill('FFF5F5F5')
        c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
        ws.merge_cells('D2:G2')

        # Row 3: column headers
        ws.row_dimensions[3].height = 20
        for col, txt in [(4, 'סניף'), (5, 'הערות'), (6, 'תאריך'), (7, 'שם')]:
            c = ws.cell(3, col, txt)
            c.font = fnt(True, '1A3C5E', 10)
            c.fill = sfill('FFE8EEF4')
            c.alignment = align('center', False)
            c.border = THIN

        td = defaultdict(list)
        for d in sheet_diffs: td[d['tutor']].append(d)
        esm  = empty_sessions_map or {}
        mdsm = missing_ds_map or {}

        data_start = row = 4
        for ti, tutor in enumerate(sorted(tutors, key=lambda x: x.split()[0])):
            diffs = td.get(tutor, [])
            by_dt = defaultdict(list)
            for d in diffs: by_dt[(d['date'], d['time'])].append(d)
            es_list  = sorted(esm.get(tutor, []),  key=lambda x: x['date'])
            mds_list = sorted(mdsm.get(tutor, []), key=lambda x: (x['date'], x['time']))
            alt = sfill('FFF7FBFF') if ti % 2 == 0 else sfill('FFFFFFFF')

            if not by_dt and not es_list and not mds_list:
                ws.row_dimensions[row].height = 18
                for col, (txt, ah) in [(4, ('~', 'center')), (5, ('ללא פערים', 'right')), (6, ('~', 'center'))]:
                    c = ws.cell(row, col, txt)
                    c.font = fnt(color='AAAAAA', italic=True); c.fill = alt; c.border = THIN
                    c.alignment = align(ah, False)
                c = ws.cell(row, 7, tutor.split()[0])
                c.font = fnt(True, '1A3C5E'); c.fill = alt; c.border = THIN
                c.alignment = align('right', False)
                row += 1
            else:
                first_row = True
                for (date, t), ds in sorted(by_dt.items()):
                    note_text = f"בשיעור של {t} - {build_note(ds)}"
                    ws.row_dimensions[row].height = note_row_height(note_text)

                    loc = next((d['location'] for d in ds if d['location']), None)
                    rf = branch_fill(loc)

                    c = ws.cell(row, 4, loc or '—')
                    c.fill = rf; c.border = THIN
                    c.alignment = align('center', False)
                    c.font = fnt(bold=True, color='333333', size=9)

                    c = ws.cell(row, 5, note_text)
                    c.fill = rf; c.border = THIN
                    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
                    c.font = fnt(size=10)

                    c = ws.cell(row, 6, date.strftime('%d.%m'))
                    c.fill = rf; c.border = THIN
                    c.alignment = align('center', False); c.font = fnt(True)

                    c = ws.cell(row, 7, tutor.split()[0] if first_row else '')
                    c.font = fnt(True, '1A3C5E') if first_row else fnt()
                    c.fill = rf; c.border = THIN
                    c.alignment = align('right', False)

                    first_row = False; row += 1

                for mds in mds_list:
                    note_text = f"שיעור חסר בדרוש תיקון — שיעור של {mds['time']} קיים בייצוא שיטס בלבד"
                    ws.row_dimensions[row].height = 18

                    c = ws.cell(row, 4, mds['location'] or '—')
                    c.fill = MISS_FILL; c.border = THIN
                    c.alignment = align('center', False)
                    c.font = fnt(bold=True, color='B71C1C', size=9)

                    c = ws.cell(row, 5, note_text)
                    c.fill = MISS_FILL; c.border = THIN
                    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
                    c.font = fnt(bold=True, color='B71C1C', size=10)

                    c = ws.cell(row, 6, mds['date'].strftime('%d.%m'))
                    c.fill = MISS_FILL; c.border = THIN
                    c.alignment = align('center', False)
                    c.font = fnt(True, 'B71C1C')

                    c = ws.cell(row, 7, tutor.split()[0] if first_row else '')
                    c.font = fnt(True, '1A3C5E') if first_row else fnt()
                    c.fill = MISS_FILL; c.border = THIN
                    c.alignment = align('right', False)

                    first_row = False; row += 1

                for es in es_list:
                    _cnt = es['count']
                    _slot_word = 'שיעור' if _cnt == 1 else 'שיעורים'
                    note_text = f"ייצוא שיטס ריק — יש {_cnt} {_slot_word} בדרוש תיקון. יש למלא בייצוא."
                    ws.row_dimensions[row].height = 18

                    c = ws.cell(row, 4, es['location'] or '—')
                    c.fill = WARN_FILL; c.border = THIN
                    c.alignment = align('center', False)
                    c.font = fnt(bold=True, color='5D4037', size=9)

                    c = ws.cell(row, 5, note_text)
                    c.fill = WARN_FILL; c.border = THIN
                    c.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
                    c.font = fnt(bold=True, color='5D4037', size=10)

                    c = ws.cell(row, 6, es['date'].strftime('%d.%m'))
                    c.fill = WARN_FILL; c.border = THIN
                    c.alignment = align('center', False)
                    c.font = fnt(True, '5D4037')

                    c = ws.cell(row, 7, tutor.split()[0] if first_row else '')
                    c.font = fnt(True, '1A3C5E') if first_row else fnt()
                    c.fill = WARN_FILL; c.border = THIN
                    c.alignment = align('right', False)

                    first_row = False; row += 1

        # Branch sidebar col B
        for bi, branch in enumerate(BRANCHES_ORDER):
            r = data_start + bi
            if r >= row: break
            bc = 'FF' + BRANCH_COLORS.get(branch, BRANCH_COLORS['default'])
            c = ws.cell(r, 2, branch)
            c.fill = sfill(bc); c.font = fnt(True, 'FFFFFF', 9)
            c.alignment = align('center', False); c.border = THIN
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height, 18)

        row += 2
        c = ws.cell(row, 4, 'פערים קודמים שטרם נסגרו')
        c.font = fnt(True, '888888', 9, True); c.fill = sfill('FFF5F5F5')
        ws.merge_cells(f'D{row}:G{row}')

    wb = Workbook(); wb.remove(wb.active)
    math_diffs = [d for d in differences if d['subject'] == 'מתמטיקה']
    eng_diffs  = [d for d in differences if d['subject'] == 'אנגלית']

    def _es_map(subj):
        m = defaultdict(list)
        for es in (empty_sessions or []):
            if es['subject'] == subj:
                m[es['tutor']].append(es)
        return m

    def _mds_map(subj):
        m = defaultdict(list)
        for mds in (missing_darush_sessions or []):
            if mds['subject'] == subj:
                m[mds['tutor']].append(mds)
        return m

    make_sheet(wb, 'מתמטיקה', math_diffs, math_tutors, _es_map('מתמטיקה'), _mds_map('מתמטיקה'))
    make_sheet(wb, 'אנגלית',  eng_diffs,  eng_tutors,  _es_map('אנגלית'),  _mds_map('אנגלית'))
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Math: {len(math_diffs)} | English: {len(eng_diffs)} | Range: {date_range_str}")


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def fix_xlsx(src, dst):
    """Fix malformed hex RGB values in xl/styles.xml (e.g. rgb="FF#RRGGBB" → rgb="FFRRGGBB")."""
    import zipfile as _zf
    with _zf.ZipFile(src, 'r') as z:
        data = {name: z.read(name) for name in z.namelist()}
    if 'xl/styles.xml' in data:
        content = data['xl/styles.xml'].decode('utf-8')
        content = re.sub(r'rgb="FF#([0-9A-Fa-f]{6})"', r'rgb="FF\1"', content)
        data['xl/styles.xml'] = content.encode('utf-8')
    with _zf.ZipFile(dst, 'w', _zf.ZIP_DEFLATED) as z:
        for name, c in data.items():
            z.writestr(name, c)


if __name__ == '__main__':
    from openpyxl import load_workbook
    fix_xlsx('/mnt/user-data/uploads/תוכנית-כיתה-2026-05-01-to-2026-05-31__1_.xlsx', '/tmp/d.xlsx')
    fix_xlsx('/mnt/user-data/uploads/שיבוץ_תלמידים_לשיעורים_-_לב_השרון__23_.xlsx', '/tmp/y.xlsx')
    wb1 = load_workbook('/tmp/d.xlsx')
    wb2 = load_workbook('/tmp/y.xlsx')
    diffs, tfm, drange, empty_sessions, missing_darush = run_comparison(wb1, wb2)

    # Filter to requested date range: 01/05/2026 — 29/05/2026
    from datetime import datetime
    start = datetime(2026, 5, 1)
    end   = datetime(2026, 5, 29)
    diffs           = [d for d in diffs           if start <= d['date'] <= end]
    empty_sessions  = [e for e in empty_sessions  if start <= e['date'] <= end]
    missing_darush  = [m for m in missing_darush  if start <= m['date'] <= end]
    drange = f"{start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}"
    print(f"Total: {len(diffs)} | {Counter(d['type'] for d in diffs)}")
    build_report(diffs, tfm, drange, '/mnt/user-data/outputs/דוח_פערים_מתרגלים.xlsx',
                 empty_sessions, missing_darush)
