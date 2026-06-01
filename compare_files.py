import difflib, re, pickle
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, time as dt_time
from collections import defaultdict, Counter

wb1 = load_workbook('/home/claude/fixed_darush.xlsx')
ws1 = wb1.active
wb2 = load_workbook('/home/claude/fixed_yitzua.xlsx')

# ── Utilities ──
def norm(n):
    if not n: return ''
    return re.sub(r'\s+',' ', re.sub(r"['\u05f3\u2019\u05f4\"]",'',str(n))).strip()

def is_time_marker(val):
    if isinstance(val, dt_time): return True
    v = str(val).strip()
    if re.match(r'^\d{1,2}:\d{2}(:\d{2})?(\s*[-–]\s*\d{1,2}:\d{2})?$', v): return True
    if re.search(r'\d{1,2}:\d{2}', v) and any(s in v for s in ['מתמטיקה','אנגלית','עברית','מדעים','היסטוריה','מרתון','ספרות']): return True
    return False

def xtime(val):
    if isinstance(val, dt_time): return f"{val.hour:02d}:{val.minute:02d}"
    m = re.search(r'(\d{1,2}:\d{2})', str(val))
    return m.group(1) if m else ''

def detect_subject(val):
    v = str(val)
    if 'מתמטיקה' in v or 'מתמטיק' in v: return 'מתמטיקה'
    if 'אנגלית' in v or 'אנגל' in v: return 'אנגלית'
    return None

def parse_date(s):
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', str(s))
    if m: return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.search(r'(\d{1,2})\.(\d{1,2})', str(s))
    if m: return datetime(2026, int(m.group(2)), int(m.group(1)))
    return None

def color_fam(rgb):
    r = str(rgb).upper() if rgb else ''
    if r in ('00000000','FFFFFFFF',''): return 'none'
    if r in ('FF00FF00','FF64FF00','FF9AC47A','FFB6D7A8'): return 'ירוק'
    if r in ('FFFF0000','FFEE0000','FFEA9999','FFF4CCCC'): return 'אדום'
    if r in ('FF4A86E8','FFCFE2F3','FFC7D2D5','FFDAE2E0'): return 'כחול'
    if r in ('FFD5A6BD','FFEAD1DC','FFFF00FF'): return 'ורוד'
    if r == 'FFFF9900': return 'כתום'
    return 'none'

def names_match(a, b):
    def clean(n):
        n = re.sub(r"['\u05f3\u2019\u05f4\"\-\(\)\[\]]",'',str(n))
        n = re.sub(r'\s+[חטיכזאבדה]\d*$','',n.strip())
        return re.sub(r'\s+',' ',n).strip()
    ac, bc = clean(a), clean(b)
    if not ac or not bc: return False
    if ac == bc: return True
    an, bn = ac.replace(' ',''), bc.replace(' ','')
    if an == bn: return True
    aw, bw = ac.split(), bc.split()
    if aw and bw and aw[0] == bw[0]:
        return difflib.SequenceMatcher(None, an, bn).ratio() >= 0.70
    return difflib.SequenceMatcher(None, an, bn).ratio() >= 0.85

def tutor_canon(raw):
    m = re.search(r'מחליפ[הת]?\s+את\s+([\u05d0-\u05ea]+)', str(raw))
    if m: return m.group(1)
    first = re.split(r'[-–\n]', str(raw))[0].strip()
    first = re.sub(r'\s+ו[\u05d0-\u05ea]+.*','',first).strip()
    return first.split()[0] if first.split() else first

def parse_slots(ws, col, start_row, first_time='', first_subject=None):
    """Returns {time: {'students': {name: color_fam}, 'subject': subj}}"""
    current_t = first_time
    current_s = first_subject
    slots = {}
    for row in range(start_row, ws.max_row+1):
        cell = ws.cell(row, col)
        if cell.value is None: continue
        val = cell.value
        vs = str(val).strip()
        if not vs: continue
        if is_time_marker(val):
            t = xtime(val)
            if t:
                current_t = t
                subj = detect_subject(vs)
                if subj: current_s = subj
        else:
            if len(vs) > 50: continue
            name = norm(vs)
            rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else '00000000'
            if current_t not in slots:
                slots[current_t] = {'students': {}, 'subject': current_s}
            slots[current_t]['students'][name] = color_fam(rgb)
    return slots

def is_junk(name):
    if not name: return True
    if re.match(r'^=', name): return True
    if re.match(r'^\d{1,2}:\d{2}', name): return True
    if re.search(r'\d{1,2}:\d{2}', name) and len(name) < 15: return True
    locs = {'רופין','צורן','גאולים','ירקון','זום','תל מונד','אלתרמן','כפר יונה','שרונה','עמיאסף','חגור','מתן'}
    return name.strip() in locs

# ── Parse דרוש תיקון ──
date_cols = {}
for col in range(1, ws1.max_column+1):
    v = ws1.cell(1,col).value
    if v: date_cols[col] = v

darush = {}
tutor_full_map = {}
date_list = sorted(date_cols.keys())
for i, sc in enumerate(date_list):
    ec = date_list[i+1] if i+1<len(date_list) else ws1.max_column+1
    date_obj = parse_date(str(date_cols[sc]))
    if not date_obj: continue
    for col in range(sc, ec):
        tv = ws1.cell(2,col).value
        if not tv: continue
        tf = norm(tv); t1 = tf.split()[0]
        tutor_full_map[t1] = tf
        ft = xtime(ws1.cell(4,col).value or '')
        for t, slot_data in parse_slots(ws1, col, 5, ft).items():
            key = (date_obj, t1, t)
            if key not in darush: darush[key] = {}
            darush[key].update(slot_data['students'])

darush_dates = set(k[0] for k in darush)

# ── Parse ייצוא שיטס ──
yitzua = []
for sn in wb2.sheetnames:
    ws = wb2[sn]
    day_cols = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(1,c).value
        if v and 'יום' in str(v): day_cols[c] = str(v)
    for col in range(1, ws.max_column+1):
        raw = str(ws.cell(2,col).value or '').strip()
        if not raw: continue
        canon = tutor_canon(raw)
        dl = next((v for c,v in sorted(day_cols.items(),reverse=True) if col>=c), '')
        date_obj = parse_date(dl)
        if not date_obj or date_obj not in darush_dates: continue
        r3 = ws.cell(3,col).value
        if r3 and isinstance(r3,str) and re.search(r'[\u05d0-\u05ea]{2,}',r3) and not re.search(r'\d+:\d+',r3):
            ft, fs = '', detect_subject(r3)
        else:
            ft = xtime(r3) if r3 else ''
            fs = detect_subject(str(r3)) if r3 else None
        for t, slot_data in parse_slots(ws, col, 4, ft, fs).items():
            if slot_data['students']:
                yitzua.append({'date':date_obj,'canon':canon,'time':t,
                               'students':slot_data['students'],
                               'subject':slot_data['subject']})

# ── Compare ──
differences = []
for sy in yitzua:
    key = (sy['date'], sy['canon'], sy['time'])
    if key not in darush:
        alts = {k:v for k,v in darush.items()
                if k[0]==sy['date'] and k[2]==sy['time'] and
                (sy['canon'] in k[1] or k[1] in sy['canon'])}
        if not alts: continue
        key = next(iter(alts))
    d_sts = darush[key]
    y_sts = sy['students']
    tutor = tutor_full_map.get(key[1], key[1])
    md, my = set(), set()
    for yn in y_sts:
        for dn in d_sts:
            if dn in md: continue
            if names_match(yn, dn):
                md.add(dn); my.add(yn)
                yc,dc = y_sts[yn], d_sts[dn]
                if yc!=dc and not(yc=='none' and dc=='none'):
                    differences.append({'tutor':tutor,'date':sy['date'],'time':sy['time'],
                                        'type':'שינוי צבע','student':yn,'color_y':yc,
                                        'color_d':dc,'subject':sy['subject']})
                break
    for yn in y_sts:
        if yn not in my and not is_junk(yn):
            differences.append({'tutor':tutor,'date':sy['date'],'time':sy['time'],
                                'type':'להוסיף','student':yn,'color_y':y_sts[yn],
                                'color_d':None,'subject':sy['subject']})
    for dn in d_sts:
        if dn not in md and not is_junk(dn):
            differences.append({'tutor':tutor,'date':sy['date'],'time':sy['time'],
                                'type':'להסיר','student':dn,'color_y':None,
                                'color_d':d_sts[dn],'subject':sy['subject']})

print(f"Total: {len(differences)} | {Counter(d['type'] for d in differences)}")
subj_cnt = Counter(d['subject'] for d in differences)
print(f"By subject: {subj_cnt}")

print("\n=== ניצן הרף ===")
for d in differences:
    if 'ניצן' in d['tutor']:
        print(f"  {d['date'].strftime('%d/%m')} {d['time']} [{d['subject']}] {d['type']} | {d['student']}")

# ── Build Excel ──
all_tutors_full = sorted(set(tutor_full_map.values()))

def fill(h): return PatternFill(fill_type='solid', fgColor=h)
def fnt(bold=False, color='222222', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
thin = Side(style='thin', color='C0C0C0')
THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
def align(h='right', wrap=True):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap, readingOrder=2)

BRANCHES = ['אלתרמן','תל מונד','גאולים','ניצני עוז','עמי אסף',
            'ירקון','צורן','זום','צור יצחק','רופין','חגור','מתן','כפר יונה']
BRANCH_COLORS = ['CCC8D3','ACA9B5','8F8E99','919398','93A0A5','A5B2BA',
                 'B7C4CA','C3D6D3','D7E6DA','C5D1BD','AEB9A3','B9B79C','CCCBAF']

def build_note(ds):
    adds = [d['student'] for d in ds if d['type']=='להוסיף']
    rems = [d['student'] for d in ds if d['type']=='להסיר']
    cols = [(d['student'],d['color_y'],d['color_d']) for d in ds if d['type']=='שינוי צבע']
    parts = []
    if adds: parts.append('להוסיף את: '+', '.join(adds))
    if rems: parts.append('להסיר את: '+', '.join(rems))
    for nm,cy,cd in cols:
        cy_s = cy if cy and cy!='none' else 'ללא צבע'
        cd_s = cd if cd and cd!='none' else 'ללא צבע'
        parts.append(f"לשנות צבע של {nm}: {cd_s} ← {cy_s}")
    return ' | '.join(parts)

def make_sheet(wb, title, sheet_diffs, tutors):
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    for col, w in zip(range(1,8), [3,14,3,16,56,10,14]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 6
    # Row 2
    ws.row_dimensions[2].height = 24
    c = ws.cell(2,2,title)
    c.font=fnt(True,'FFFFFF',12); c.fill=fill('1A3C5E'); c.alignment=align('center',False)
    shorts = sorted(t.split()[0] for t in tutors)
    c = ws.cell(2,4,'מדריכים שנבדקו: '+', '.join(shorts)+'.')
    c.font=fnt(size=9,color='555555'); c.fill=fill('F5F5F5'); c.alignment=align('right')
    ws.merge_cells('D2:G2')
    # Row 3
    ws.row_dimensions[3].height = 20
    for col,txt in [(4,'סניף'),(5,'הערות'),(6,'תאריך'),(7,'שם')]:
        c=ws.cell(3,col,txt)
        c.font=fnt(True,'1A3C5E',10); c.fill=fill('E8EEF4')
        c.alignment=align('center',False); c.border=THIN

    td = defaultdict(list)
    for d in sheet_diffs: td[d['tutor']].append(d)

    data_start = row = 4
    for ti, tutor in enumerate(sorted(tutors, key=lambda x: x.split()[0])):
        diffs = td.get(tutor,[])
        by_dt = defaultdict(list)
        for d in diffs: by_dt[(d['date'],d['time'])].append(d)
        alt = fill('F7FBFF') if ti%2==0 else fill('FFFFFF')
        if not by_dt:
            ws.row_dimensions[row].height = 18
            for col,(txt,ah) in [(4,('~','center')),(5,('ללא פערים','right')),(6,('~','center'))]:
                c=ws.cell(row,col,txt)
                c.font=fnt(color='AAAAAA',italic=True); c.fill=alt; c.border=THIN; c.alignment=align(ah,False)
            c=ws.cell(row,7,tutor.split()[0])
            c.font=fnt(True,'1A3C5E'); c.fill=alt; c.border=THIN; c.alignment=align('right',False)
            row+=1
        else:
            first_row=True
            for (date,t),ds in sorted(by_dt.items()):
                ws.row_dimensions[row].height=32
                ha=any(d['type']=='להוסיף' for d in ds)
                hr=any(d['type']=='להסיר' for d in ds)
                rf=fill('FFF3E0') if ha and hr else fill('E8F5E9') if ha else fill('FFEBEE') if hr else fill('FFFDE7')
                note=f"בשיעור של {t} - {build_note(ds)}"
                c=ws.cell(row,4,'—'); c.fill=rf; c.border=THIN; c.alignment=align('center',False); c.font=fnt(color='AAAAAA')
                c=ws.cell(row,5,note); c.fill=rf; c.border=THIN; c.alignment=align('right',True); c.font=fnt(size=10)
                c=ws.cell(row,6,date.strftime('%d.%m')); c.fill=rf; c.border=THIN; c.alignment=align('center',False); c.font=fnt(True)
                c=ws.cell(row,7,tutor.split()[0] if first_row else '')
                c.font=fnt(True,'1A3C5E') if first_row else fnt()
                c.fill=rf; c.border=THIN; c.alignment=align('right',False)
                first_row=False; row+=1
    for bi,(branch,bc) in enumerate(zip(BRANCHES,BRANCH_COLORS)):
        r=data_start+bi
        if r>=row: break
        c=ws.cell(r,2,branch)
        c.fill=fill(bc); c.font=fnt(True,'FFFFFF',9)
        c.alignment=align('center',False); c.border=THIN
        ws.row_dimensions[r].height=max(ws.row_dimensions[r].height,18)
    row+=2
    c=ws.cell(row,4,'פערים קודמים שטרם נסגרו')
    c.font=fnt(True,'888888',9,True); c.fill=fill('F5F5F5')
    ws.merge_cells(f'D{row}:G{row}')

wb = Workbook()
wb.remove(wb.active)

math_diffs = [d for d in differences if d['subject']=='מתמטיקה']
eng_diffs  = [d for d in differences if d['subject']=='אנגלית']
gen_diffs  = [d for d in differences if d['subject'] is None]

# מתמטיקה: math diffs + unclassified general
# אנגלית: english diffs only
make_sheet(wb, 'מתמטיקה', math_diffs + gen_diffs, all_tutors_full)
make_sheet(wb, 'אנגלית',  eng_diffs,               all_tutors_full)

output='/mnt/user-data/outputs/דוח_פערים_מתרגלים.xlsx'
wb.save(output)
print(f"\nSaved: {output}")
print(f"Math+gen: {len(math_diffs)+len(gen_diffs)} | English: {len(eng_diffs)}")
