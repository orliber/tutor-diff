#!/usr/bin/env python3
"""
Tutor Diff — Desktop GUI
PyQt6 wrapper around build_excel.py.
"""

import sys
import os
import re
import shutil
import tempfile
import subprocess
from datetime import datetime
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFrame, QDialog, QScrollArea,
    QSizePolicy, QProgressBar, QGraphicsDropShadowEffect,
    QAbstractButton, QCalendarWidget,
)
from PyQt6.QtCore import (
    Qt, QDate, QThread, pyqtSignal, QSettings, QSize, QStandardPaths,
    QVariantAnimation, QEasingCurve, QLocale, QTimer, QPoint,
    QEvent, QRect,
)
from PyQt6.QtGui import QFont, QDragEnterEvent, QDropEvent, QColor, QPainter, QShortcut, QKeySequence

from openpyxl import load_workbook
from build_excel import run_comparison, build_report, fix_xlsx

# ── Palette ───────────────────────────────────────────────────────────────────
BG        = '#eef2f7'
CARD      = '#ffffff'
PRIMARY   = '#1e3a5f'
PRIMARY_H = '#2d5a8e'
ACCENT    = '#3b82f6'
SUCCESS   = '#059669'
SUCCESS_H = '#047857'
MUTED     = '#64748b'
BORDER    = '#dde3ea'
TEXT      = '#1e293b'


# ── File-type detection ───────────────────────────────────────────────────────

def _detect_file_type(path: str) -> 'str | None':
    """Return 'darush', 'yitzua', or None by inspecting row 1 of the workbook."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            vals = [str(v or '') for v in first_row[:60]]
            if any('יום' in v for v in vals):
                wb.close(); return 'yitzua'
            if any(re.search(r'\d{1,2}[./]\d{1,2}', v) for v in vals):
                wb.close(); return 'darush'
        wb.close()
    except Exception:
        pass
    return None


# ── Styled calendar popup ─────────────────────────────────────────────────────

class StyledCalendar(QCalendarWidget):
    """Beautiful calendar widget — used via QDateEdit.setCalendarWidget()."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setGridVisible(False)
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)
        self.setHorizontalHeaderFormat(QCalendarWidget.HorizontalHeaderFormat.SingleLetterDayNames)
        self.setFirstDayOfWeek(Qt.DayOfWeek.Sunday)
        self.setLocale(QLocale(QLocale.Language.Hebrew, QLocale.Country.Israel))
        self.setMinimumWidth(290)
        self.setMinimumHeight(270)
        self._hovered:    QDate | None = None
        self._cell_rects: dict        = {}
        self._apply_style()
        QTimer.singleShot(0, self._style_nav_buttons)
        QTimer.singleShot(0, self._setup_hover)

    def _setup_hover(self):
        from PyQt6.QtWidgets import QAbstractItemView
        view = self.findChild(QAbstractItemView)
        if view:
            view.setMouseTracking(True)
            view.installEventFilter(self)

    def eventFilter(self, obj, event):
        from PyQt6.QtWidgets import QAbstractItemView
        if isinstance(obj, QAbstractItemView):
            t = event.type()
            if t == QEvent.Type.MouseMove:
                new_h = self._date_at(event.pos())
                if new_h != self._hovered:
                    self._hovered = new_h
                    self.updateCells()
            elif t == QEvent.Type.Leave:
                if self._hovered is not None:
                    self._hovered = None
                    self.updateCells()
        return super().eventFilter(obj, event)

    def _date_at(self, pos) -> 'QDate | None':
        for date, rect in self._cell_rects.items():
            if rect.contains(pos):
                return date
        return None

    def _apply_style(self):
        self.setStyleSheet(f"""
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {PRIMARY}, stop:1 {PRIMARY_H});
                min-height: 50px;
                padding: 0 8px;
            }}
            QCalendarWidget QToolButton {{
                color: white; background: transparent;
                border: none; font-size: 14px; font-weight: bold;
                padding: 4px 10px; border-radius: 7px;
            }}
            QCalendarWidget QToolButton:hover {{
                background: rgba(255,255,255,0.18);
            }}
            QCalendarWidget QToolButton::menu-indicator {{
                image: none; width: 0;
            }}
            QCalendarWidget QSpinBox {{
                color: white;
                background: rgba(255,255,255,0.15);
                border: 1px solid rgba(255,255,255,0.25);
                border-radius: 5px; padding: 2px 6px;
                font-size: 13px; font-weight: bold;
            }}
            QCalendarWidget QSpinBox::up-button,
            QCalendarWidget QSpinBox::down-button {{ width: 0; height: 0; }}
            QCalendarWidget QAbstractItemView:enabled {{
                color: {TEXT}; background: white;
                selection-background-color: transparent;
                selection-color: white;
                outline: none; font-size: 13px;
            }}
            QCalendarWidget QAbstractItemView:disabled {{ color: #cbd5e1; }}
            QCalendarWidget QWidget {{ alternate-background-color: white; }}
        """)

    def _style_nav_buttons(self):
        for name, arrow in (('qt_calendar_prevmonth', '‹'),
                             ('qt_calendar_nextmonth', '›')):
            btn = self.findChild(QPushButton, name)
            if btn:
                btn.setText(arrow)
                btn.setStyleSheet("""
                    QPushButton {
                        color: white;
                        background: rgba(255,255,255,0.15);
                        border: none; border-radius: 7px;
                        font-size: 20px; font-weight: 300;
                        min-width: 32px; max-width: 32px;
                        min-height: 32px; max-height: 32px;
                    }
                    QPushButton:hover  { background: rgba(255,255,255,0.28); }
                    QPushButton:pressed { background: rgba(255,255,255,0.08); }
                """)

    def updateCells(self):
        self._cell_rects.clear()  # stale rects from previous months would cause wrong hover
        super().updateCells()

    def paintCell(self, painter, rect, date):
        if not date.isValid():
            return

        # Store rect for hover hit-testing (cleared on each updateCells cycle)
        self._cell_rects[date] = QRect(
            int(rect.x()), int(rect.y()),
            int(rect.width()), int(rect.height())
        )

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setPen(Qt.PenStyle.NoPen)

        today    = QDate.currentDate()
        selected = self.selectedDate()
        other_mo = date.month() != self.monthShown()
        hovered  = date == self._hovered and not other_mo

        d = min(rect.width(), rect.height()) - 8
        x = rect.center().x() - d / 2
        y = rect.center().y() - d / 2

        if date == selected:
            painter.setBrush(QColor(ACCENT))
            painter.drawEllipse(x, y, d, d)
            painter.setPen(QColor('#ffffff'))
            f = QFont(); f.setPointSize(12); f.setBold(True)
        elif date == today:
            painter.setBrush(QColor('#dbeafe'))
            painter.drawEllipse(x, y, d, d)
            painter.setPen(QColor(ACCENT))
            f = QFont(); f.setPointSize(12); f.setBold(True)
        elif hovered:
            painter.setBrush(QColor('#e8edf2'))
            painter.drawEllipse(x, y, d, d)
            painter.setPen(QColor(TEXT))
            f = QFont(); f.setPointSize(12)
        else:
            color = '#b8c8d8' if other_mo else ('#64748b' if date.dayOfWeek() == 7 else TEXT)
            painter.setPen(QColor(color))
            f = QFont(); f.setPointSize(12)

        painter.setFont(f)
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, str(date.day()))
        painter.restore()


# ── Date picker button ────────────────────────────────────────────────────────

class DatePickerButton(QPushButton):
    """Button that shows the selected date and opens StyledCalendar on click."""
    dateChanged = pyqtSignal(QDate)

    def __init__(self, initial: QDate, parent=None):
        super().__init__(parent)
        self._date = initial
        self._cal  = StyledCalendar()
        self._cal.setWindowFlags(
            Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint
        )
        self._cal.setSelectedDate(initial)
        self._cal.clicked.connect(self._pick)
        self.clicked.connect(self._open)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(36)
        self.setMinimumWidth(140)
        self.setStyleSheet(f"""
            QPushButton {{
                border: 1.5px solid {BORDER}; border-radius: 8px;
                padding: 5px 14px; background: #f8fafc;
                color: {TEXT}; font-size: 13px; text-align: center;
            }}
            QPushButton:hover {{
                background: white; border-color: {ACCENT};
            }}
        """)
        self._refresh()

    def _refresh(self):
        self.setText('📅  ' + self._date.toString('dd.MM.yyyy'))

    def _open(self):
        self._cal.setSelectedDate(self._date)
        # Position below the button
        pos = self.mapToGlobal(QPoint(0, self.height() + 4))
        self._cal.move(pos)
        self._cal.show()
        self._cal.raise_()

    def _pick(self, date: QDate):
        self._date = date
        self._refresh()
        self._cal.hide()
        self.dateChanged.emit(date)

    def date(self) -> QDate:
        return self._date


def _shadow(blur=20, y=5, alpha=30):
    e = QGraphicsDropShadowEffect()
    e.setBlurRadius(blur)
    e.setXOffset(0)
    e.setYOffset(y)
    e.setColor(QColor(0, 0, 0, alpha))
    return e


def _alert(parent, title: str, msg: str, kind: str = 'warning'):
    """Styled alert — replaces QMessageBox to avoid the white-text-on-dark-theme bug."""
    dlg = QDialog(parent)
    dlg.setWindowTitle(title)
    dlg.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
    dlg.setStyleSheet(f'background: {CARD};')
    dlg.setMinimumWidth(360)
    dlg.setMaximumWidth(520)

    lay = QVBoxLayout(dlg)
    lay.setContentsMargins(24, 20, 24, 20)
    lay.setSpacing(16)

    lbl = QLabel(msg)
    lbl.setWordWrap(True)
    lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
    if kind == 'error':
        lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        lbl.setStyleSheet(f'color: {TEXT}; font-size: 11px; font-family: monospace; background: transparent;')
    else:
        lbl.setStyleSheet(f'color: {TEXT}; font-size: 13px; background: transparent;')
    lay.addWidget(lbl)

    btn_color  = '#d97706' if kind == 'warning' else '#dc2626'
    btn_hcolor = '#b45309' if kind == 'warning' else '#b91c1c'
    ok_btn = QPushButton('אישור')
    ok_btn.setCursor(Qt.CursorShape.PointingHandCursor)
    ok_btn.setStyleSheet(f"""
        QPushButton {{
            background: {btn_color}; color: white; border-radius: 8px;
            padding: 8px 28px; font-size: 13px; font-weight: bold; border: none;
        }}
        QPushButton:hover {{ background: {btn_hcolor}; }}
    """)
    ok_btn.clicked.connect(dlg.accept)
    lay.addWidget(ok_btn, 0, Qt.AlignmentFlag.AlignCenter)
    dlg.exec()


# ── Animated iOS-style toggle ─────────────────────────────────────────────────

class ToggleSwitch(QAbstractButton):
    """Smooth animated toggle — uses QAbstractButton for reliable click handling."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setFixedSize(46, 26)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._pos = 3.0
        self._anim = QVariantAnimation(self)
        self._anim.setDuration(150)
        self._anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        self._anim.valueChanged.connect(self._update)
        self.toggled.connect(self._kick)

    def _update(self, v):
        self._pos = float(v)
        self.update()

    def _kick(self, checked: bool):
        self._anim.setStartValue(float(self._pos))
        self._anim.setEndValue(22.0 if checked else 3.0)
        self._anim.start()

    def paintEvent(self, _event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setPen(Qt.PenStyle.NoPen)
        p.setBrush(QColor(ACCENT if self.isChecked() else '#c8d5e0'))
        p.drawRoundedRect(0, 0, 46, 26, 13, 13)
        p.setBrush(QColor('#ffffff'))
        p.drawEllipse(int(self._pos), 3, 20, 20)
        p.end()

    def sizeHint(self):
        return QSize(46, 26)


# ── Drop zone ─────────────────────────────────────────────────────────────────

class DropZone(QFrame):
    file_changed = pyqtSignal(str)
    cleared      = pyqtSignal()

    def __init__(self, title: str, subtitle: str, parent=None):
        super().__init__(parent)
        self.file_path: str | None = None
        self.setAcceptDrops(True)
        self.setMinimumHeight(172)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.setGraphicsEffect(_shadow())
        self._build(title, subtitle)
        self._apply_style('empty')

    def _build(self, title, subtitle):
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.setSpacing(5)
        lay.setContentsMargins(16, 16, 16, 14)

        self._icon = QLabel('📂')
        self._icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._icon.setFont(QFont('', 32))

        self._title = QLabel(title)
        self._title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        f = QFont(); f.setBold(True); f.setPointSize(14)
        self._title.setFont(f)
        self._title.setStyleSheet(f'color: {TEXT}; background: transparent;')

        self._sub = QLabel(subtitle)
        self._sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._sub.setStyleSheet(f'color: {MUTED}; font-size: 11px; background: transparent;')

        self._file = QLabel('')
        self._file.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._file.setStyleSheet(f'color: {ACCENT}; font-size: 11px; font-weight: bold; background: transparent;')
        self._file.setWordWrap(True)

        self._meta = QLabel('')
        self._meta.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._meta.setStyleSheet('color: #9aa5b1; font-size: 10px; background: transparent;')

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        btn_row.addStretch()

        self._choose_btn = QPushButton('בחר קובץ')
        self._choose_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._choose_btn.setStyleSheet(f"""
            QPushButton {{
                background: #f1f5f9; border: 1px solid {BORDER};
                border-radius: 6px; padding: 5px 16px; font-size: 12px; color: {TEXT};
            }}
            QPushButton:hover {{ background: #e2e8f0; }}
        """)
        self._choose_btn.clicked.connect(self._choose)

        self._clear_btn = QPushButton('✕')
        self._clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._clear_btn.setFixedSize(28, 28)
        self._clear_btn.setToolTip('נקה קובץ')
        self._clear_btn.setStyleSheet("""
            QPushButton {
                background: #fee2e2; border: 1px solid #fca5a5;
                border-radius: 6px; font-size: 11px; color: #dc2626;
            }
            QPushButton:hover { background: #fca5a5; }
        """)
        self._clear_btn.clicked.connect(self._clear)
        self._clear_btn.setVisible(False)

        btn_row.addWidget(self._choose_btn)
        btn_row.addWidget(self._clear_btn)
        btn_row.addStretch()

        for w in (self._icon, self._title, self._sub, self._file, self._meta):
            lay.addWidget(w)
        lay.addLayout(btn_row)

    def _apply_style(self, state: str):
        styles = {
            'empty':  f'background:{CARD}; border:2px dashed #c0cdd9; border-radius:14px;',
            'filled': f'background:#f0f7ff; border:2px solid {ACCENT}; border-radius:14px;',
            'hover':  f'background:#dbeafe; border:2px dashed {ACCENT}; border-radius:14px;',
        }
        self.setStyleSheet(f'DropZone {{ {styles[state]} }}')
        self._icon.setText('✅' if state == 'filled' else '📂')

    def set_file(self, path: str):
        self.file_path = path
        name = Path(path).name
        self._file.setText(name if len(name) <= 36 else name[:33] + '…')
        self._file.setToolTip(path)
        try:
            kb   = os.path.getsize(path) // 1024
            date = datetime.fromtimestamp(os.path.getmtime(path)).strftime('%d.%m.%Y')
            self._meta.setText(f'{kb:,} KB  ·  {date}')
        except Exception:
            self._meta.setText('')
        self._clear_btn.setVisible(True)
        self._apply_style('filled')
        self.file_changed.emit(path)

    def _clear(self):
        self.file_path = None
        self._file.setText('')
        self._file.setToolTip('')
        self._meta.setText('')
        self._clear_btn.setVisible(False)
        self._apply_style('empty')
        self.cleared.emit()

    def _choose(self):
        from PyQt6.QtWidgets import QFileDialog
        start = str(Path(self.file_path).parent) if self.file_path else os.path.expanduser('~')
        path, _ = QFileDialog.getOpenFileName(self, 'בחר קובץ Excel', start, 'Excel (*.xlsx *.xls)')
        if path:
            self.set_file(path)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0].toLocalFile().lower()
            if url.endswith(('.xlsx', '.xls')):
                event.acceptProposedAction()
                self._apply_style('hover')

    def dragLeaveEvent(self, event):
        self._apply_style('filled' if self.file_path else 'empty')

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls:
            self.set_file(urls[0].toLocalFile())


# ── Date range card ───────────────────────────────────────────────────────────

class DateRangeCard(QFrame):
    """Clean card with toggle + inline date pickers. No QGroupBox."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            DateRangeCard {{
                background: {CARD};
                border: 1px solid {BORDER};
                border-radius: 14px;
            }}
        """)
        self.setGraphicsEffect(_shadow(14, 3, 20))
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(18, 14, 18, 14)
        root.setSpacing(12)

        # ── Header row ────────────────────────────────────────────────
        header = QHBoxLayout()
        header.setSpacing(10)

        icon_lbl = QLabel('📅')
        icon_lbl.setStyleSheet('font-size: 16px; background: transparent;')

        text_lbl = QLabel('סינון לפי טווח תאריכים')
        text_lbl.setStyleSheet(f'color: {TEXT}; font-size: 13px; font-weight: 600; background: transparent;')

        hint_lbl = QLabel('(אופציונלי)')
        hint_lbl.setStyleSheet(f'color: {MUTED}; font-size: 11px; background: transparent;')

        self.toggle = ToggleSwitch()
        self.toggle.toggled.connect(self._on_toggle)

        header.addWidget(icon_lbl)
        header.addWidget(text_lbl)
        header.addWidget(hint_lbl)
        header.addStretch()
        header.addWidget(self.toggle)
        root.addLayout(header)

        # ── Date pickers row (hidden by default) ───────────────────────
        self._pickers = QWidget()
        self._pickers.setStyleSheet('background: transparent;')
        prow = QHBoxLayout(self._pickers)
        prow.setContentsMargins(0, 0, 0, 0)
        prow.setSpacing(10)

        from_lbl = QLabel('מ:')
        from_lbl.setStyleSheet(f'color:{MUTED}; font-size:12px; background:transparent;')
        self.date_from = DatePickerButton(QDate.currentDate().addDays(-30))

        sep = QLabel('—')
        sep.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sep.setStyleSheet(f'color:{MUTED}; font-size:14px; background:transparent;')

        to_lbl = QLabel('עד:')
        to_lbl.setStyleSheet(f'color:{MUTED}; font-size:12px; background:transparent;')
        self.date_to = DatePickerButton(QDate.currentDate())

        prow.addWidget(from_lbl)
        prow.addWidget(self.date_from)
        prow.addWidget(sep)
        prow.addWidget(to_lbl)
        prow.addWidget(self.date_to)
        prow.addStretch()

        self._pickers.setVisible(False)
        root.addWidget(self._pickers)

    def _on_toggle(self, checked: bool):
        self._pickers.setVisible(checked)

    def is_active(self):
        return self.toggle.isChecked()

    def get_start(self):
        d = self.date_from.date()
        return datetime(d.year(), d.month(), d.day())

    def get_end(self):
        d = self.date_to.date()
        return datetime(d.year(), d.month(), d.day())


# ── Results card ──────────────────────────────────────────────────────────────

class ResultsCard(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setVisible(False)
        self._output_path: str | None = None
        self.setStyleSheet("""
            ResultsCard {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #ecfdf5, stop:1 #f0fdf4);
                border: 1.5px solid #6ee7b7;
                border-radius: 14px;
            }
        """)
        self.setGraphicsEffect(_shadow(16, 4, 25))

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        self._stats = QLabel()
        self._stats.setAlignment(Qt.AlignmentFlag.AlignCenter)
        f = QFont(); f.setPointSize(13); f.setBold(True)
        self._stats.setFont(f)
        self._stats.setStyleSheet('color: #065f46; background: transparent;')

        self._open_btn = QPushButton('  פתח את הדוח  📂')
        self._open_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._open_btn.setMinimumHeight(40)
        self._open_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {SUCCESS}, stop:1 #10b981);
                color: white; border-radius: 9px;
                padding: 6px 28px; font-size: 14px; font-weight: bold;
            }}
            QPushButton:hover   {{ background: {SUCCESS_H}; }}
            QPushButton:pressed {{ background: #065f46; }}
        """)
        self._open_btn.clicked.connect(self._open)

        lay.addWidget(self._stats)
        lay.addWidget(self._open_btn, 0, Qt.AlignmentFlag.AlignCenter)

    def show_results(self, path: str, math_count: int, eng_count: int,
                     mark_count: int = 0, warnings_count: int = 0):
        self._output_path = path
        total = math_count + eng_count
        lines = [f'✓   נמצאו {total} פערים   ·   מתמטיקה: {math_count}   |   אנגלית: {eng_count}']
        extra = []
        if mark_count:
            extra.append(f'לסמן באדום: {mark_count}')
        if warnings_count:
            extra.append(f'אזהרות: {warnings_count}')
        if extra:
            lines.append('   ·   '.join(extra))
        self._stats.setText('\n'.join(lines))
        self.setVisible(True)

    def hide_results(self):
        self.setVisible(False)
        self._output_path = None

    def _open(self):
        if self._output_path:
            subprocess.run(['open', self._output_path], check=False)


# ── Worker ────────────────────────────────────────────────────────────────────

class Worker(QThread):
    finished = pyqtSignal(str, int, int, int, int, str)   # tmp_path, math, eng, mark, warnings, drange
    error    = pyqtSignal(str)
    status   = pyqtSignal(str)
    progress = pyqtSignal(int)

    def __init__(self, darush, yitzua, use_filter, date_start, date_end, output):
        super().__init__()
        self.darush     = darush
        self.yitzua     = yitzua
        self.use_filter = use_filter
        self.date_start = date_start
        self.date_end   = date_end
        self.output     = output

    def run(self):
        tmp     = tempfile.mkdtemp()
        d_fixed = os.path.join(tmp, 'd.xlsx')
        y_fixed = os.path.join(tmp, 'y.xlsx')
        try:
            self.status.emit('מתקן קבצים…');   self.progress.emit(10)
            fix_xlsx(self.darush, d_fixed)
            fix_xlsx(self.yitzua, y_fixed)

            self.status.emit('טוען קבצים…');   self.progress.emit(25)
            wb1 = load_workbook(d_fixed)
            wb2 = load_workbook(y_fixed)

            # ── בדיקות מבנה בסיסיות ────────────────────────────────────
            ws1 = wb1.active
            if ws1 is None or ws1.max_row < 3:
                self.error.emit('קובץ "דרוש תיקון" נראה ריק.\nוודא שבחרת את הקובץ הנכון.')
                return

            has_dates = any(
                ws1.cell(1, c).value and
                re.search(r'\d{1,2}[./]\d{1,2}', str(ws1.cell(1, c).value))
                for c in range(1, min(ws1.max_column + 1, 30))
            )
            if not has_dates:
                self.error.emit(
                    'קובץ "דרוש תיקון" לא נראה כתוכנית שיעורים.\n'
                    'שורה 1 צריכה להכיל תאריכים (למשל 01.05.2026).'
                )
                return

            has_yom = any(
                'יום' in str(wb2[sn].cell(1, c).value or '')
                for sn in wb2.sheetnames
                for c in range(1, min(wb2[sn].max_column + 1, 60))
            )
            if not has_yom:
                self.error.emit(
                    'קובץ "ייצוא שיטס" לא נראה כייצוא תקין.\n'
                    'שורה 1 צריכה להכיל "יום ראשון", "יום שני" וכו\'.'
                )
                return
            # ────────────────────────────────────────────────────────────

            self.status.emit('משווה…');         self.progress.emit(50)
            diffs, tfm, drange, empty_sessions, missing_darush = run_comparison(wb1, wb2)

            if self.use_filter:
                diffs  = [d for d in diffs if self.date_start <= d['date'] <= self.date_end]
                empty_sessions  = [es  for es  in empty_sessions
                                   if self.date_start <= es['date']  <= self.date_end]
                missing_darush  = [mds for mds in missing_darush
                                   if self.date_start <= mds['date'] <= self.date_end]
                drange = (f"{self.date_start.strftime('%d.%m.%Y')} — "
                          f"{self.date_end.strftime('%d.%m.%Y')}")

            if not diffs and not empty_sessions and not missing_darush:
                if self.use_filter:
                    self.error.emit(
                        f'לא נמצאו פערים בין '
                        f'{self.date_start.strftime("%d.%m.%Y")} ל-{self.date_end.strftime("%d.%m.%Y")}.\n'
                        'נסה להרחיב את טווח התאריכים או לבטל את הסינון.'
                    )
                else:
                    self.error.emit('לא נמצאו פערים — הקבצים נראים תואמים לחלוטין.')
                return

            self.status.emit('בונה דוח…');     self.progress.emit(80)
            build_report(diffs, tfm, drange, self.output, empty_sessions, missing_darush)

            math_count     = len([d for d in diffs if d['subject'] == 'מתמטיקה'])
            eng_count      = len([d for d in diffs if d['subject'] == 'אנגלית'])
            mark_count     = len([d for d in diffs if d['type'] == 'לסמן באדום'])
            warnings_count = len(empty_sessions) + len(missing_darush)
            self.progress.emit(100)
            self.finished.emit(self.output, math_count, eng_count, mark_count, warnings_count, drange)

        except Exception as exc:
            import traceback
            self.error.emit(f'שגיאה בעיבוד הקבצים:\n{exc}\n\n{traceback.format_exc()}')
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


# ── History dialog ────────────────────────────────────────────────────────────

class HistoryDialog(QDialog):
    def __init__(self, history: list, settings: QSettings, parent=None):
        super().__init__(parent)
        self._history  = list(history)
        self._settings = settings

        self.setWindowTitle('היסטוריית ריצות')
        self.setMinimumWidth(540)
        self.setMinimumHeight(340)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setStyleSheet(f'background: {BG};')

        self._lay = QVBoxLayout(self)
        self._lay.setContentsMargins(22, 18, 22, 20)
        self._lay.setSpacing(12)

        hdr = QLabel('📋   היסטוריית ריצות')
        f = QFont(); f.setPointSize(15); f.setBold(True)
        hdr.setFont(f)
        hdr.setStyleSheet(f'color: {TEXT};')
        self._lay.addWidget(hdr)

        self._build_list()

        # ── Footer ────────────────────────────────────────────────────────
        footer = QHBoxLayout()
        footer.setSpacing(10)

        self._clear_all_btn = QPushButton('🗑   מחק הכל')
        self._clear_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._clear_all_btn.setVisible(bool(self._history))
        self._clear_all_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; border: 1px solid #fca5a5;
                border-radius: 8px; padding: 8px 18px;
                font-size: 13px; color: #dc2626;
            }}
            QPushButton:hover {{ background: #fee2e2; }}
        """)
        self._clear_all_btn.clicked.connect(self._delete_all)

        close_btn = QPushButton('סגור')
        close_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        close_btn.clicked.connect(self.accept)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: {PRIMARY}; color: white; border-radius: 8px;
                padding: 8px 28px; font-size: 13px; font-weight: bold;
            }}
            QPushButton:hover {{ background: {PRIMARY_H}; }}
        """)

        footer.addWidget(self._clear_all_btn)
        footer.addStretch()
        footer.addWidget(close_btn)
        self._lay.addLayout(footer)

    def _build_list(self):
        if not self._history:
            self._scroll = None
            empty = QLabel('אין ריצות קודמות עדיין.')
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty.setStyleSheet(f'color: {MUTED}; font-size: 13px; padding: 30px;')
            self._lay.insertWidget(1, empty)
            self._lay.insertStretch(2)
            return

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet('QScrollArea { border: none; background: transparent; }')
        container = QWidget()
        container.setStyleSheet('background: transparent;')
        self._vlay = QVBoxLayout(container)
        self._vlay.setSpacing(8)
        self._vlay.setContentsMargins(2, 2, 2, 2)

        for entry in self._history:
            self._vlay.addWidget(self._make_card(entry))

        self._vlay.addStretch()
        self._scroll.setWidget(container)
        self._lay.insertWidget(1, self._scroll)

    def _make_card(self, entry: dict) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: {CARD}; border: 1px solid {BORDER};
                border-radius: 10px;
            }}
        """)
        outer = QVBoxLayout(card)
        outer.setContentsMargins(14, 10, 10, 8)
        outer.setSpacing(3)

        # ── Main row ──────────────────────────────────────────────────────
        crow = QHBoxLayout()
        crow.setSpacing(6)

        total = entry.get('math_count', 0) + entry.get('eng_count', 0)
        parts = [f'{total} פערים']
        if entry.get('mark_count'):
            parts.append(f"לסמן באדום: {entry['mark_count']}")
        if entry.get('warnings_count'):
            parts.append(f"אזהרות: {entry['warnings_count']}")

        stat_lbl = QLabel('   ·   '.join(parts))
        sf = QFont(); sf.setBold(True); sf.setPointSize(12)
        stat_lbl.setFont(sf)
        stat_lbl.setStyleSheet(f'color: {TEXT}; background: transparent;')

        ts_lbl = QLabel(entry.get('timestamp', ''))
        ts_lbl.setStyleSheet(f'color: {MUTED}; font-size: 11px; background: transparent;')

        crow.addWidget(stat_lbl)
        crow.addStretch()

        for icon, tip, path_key, is_dir in [
            ('📄', 'פתח דוח פערים',                             'report_path', False),
            ('📂', 'פתח תיקיית הריצה (דרוש תיקון + ייצוא שיטס)', 'run_dir',     True),
        ]:
            p = entry.get(path_key, '')
            exists = os.path.isdir(p) if is_dir else os.path.exists(p)
            if p and exists:
                crow.addWidget(self._open_btn(icon, tip, p))

        crow.addWidget(ts_lbl)

        del_btn = QPushButton('✕')
        del_btn.setToolTip('הסר מההיסטוריה')
        del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        del_btn.setFixedSize(22, 22)
        del_btn.setStyleSheet("""
            QPushButton {
                background: transparent; border: none;
                color: #94a3b8; font-size: 13px; font-weight: bold;
            }
            QPushButton:hover { color: #dc2626; }
        """)
        del_btn.clicked.connect(lambda _, e=entry, c=card: self._delete_entry(e, c))
        crow.addWidget(del_btn)

        outer.addLayout(crow)

        # ── Date range subtitle ───────────────────────────────────────────
        drange = entry.get('drange', '')
        if drange:
            dr_lbl = QLabel(drange)
            dr_lbl.setStyleSheet(f'color: {MUTED}; font-size: 11px; background: transparent;')
            outer.addWidget(dr_lbl)

        return card

    @staticmethod
    def _open_btn(icon: str, tip: str, path: str) -> QPushButton:
        btn = QPushButton(icon)
        btn.setToolTip(tip)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setFixedSize(30, 30)
        btn.setStyleSheet(f"""
            QPushButton {{
                background: #f1f5f9; border: 1px solid {BORDER};
                border-radius: 7px; font-size: 15px;
            }}
            QPushButton:hover {{ background: #dbeafe; border-color: {ACCENT}; }}
        """)
        btn.clicked.connect(lambda _, p=path: subprocess.run(['open', p], check=False))
        return btn

    def _delete_entry(self, entry: dict, card: QFrame):
        import json, shutil
        run_dir = entry.get('run_dir', '')
        if run_dir and os.path.isdir(run_dir):
            try: shutil.rmtree(run_dir)
            except Exception: pass
        try:
            self._history.remove(entry)
        except ValueError:
            pass
        self._settings.setValue('run_history', json.dumps(self._history))
        card.hide()
        card.setMaximumHeight(0)
        if not self._history and self._clear_all_btn:
            self._clear_all_btn.setVisible(False)

    def _delete_all(self):
        import json, shutil

        dlg = QDialog(self)
        dlg.setWindowTitle('אישור מחיקה')
        dlg.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dlg.setStyleSheet(f'background: {CARD};')
        dlg.setFixedWidth(340)

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(26, 22, 26, 22)
        lay.setSpacing(20)

        msg = QLabel('למחוק את כל ההיסטוריה והקבצים השמורים?')
        msg.setWordWrap(True)
        msg.setAlignment(Qt.AlignmentFlag.AlignCenter)
        msg.setStyleSheet(f'color: {TEXT}; font-size: 14px; background: transparent;')
        lay.addWidget(msg)

        row = QHBoxLayout()
        row.setSpacing(10)

        yes_btn = QPushButton('כן, מחק הכל')
        yes_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        yes_btn.setStyleSheet("""
            QPushButton {
                background: #dc2626; color: white; border-radius: 8px;
                padding: 8px 20px; font-size: 13px; font-weight: bold; border: none;
            }
            QPushButton:hover { background: #b91c1c; }
        """)
        yes_btn.clicked.connect(dlg.accept)

        no_btn = QPushButton('ביטול')
        no_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        no_btn.setStyleSheet(f"""
            QPushButton {{
                background: #f1f5f9; border: 1px solid {BORDER};
                border-radius: 8px; padding: 8px 20px;
                font-size: 13px; color: {TEXT};
            }}
            QPushButton:hover {{ background: #e2e8f0; }}
        """)
        no_btn.clicked.connect(dlg.reject)

        row.addWidget(yes_btn)
        row.addWidget(no_btn)
        lay.addLayout(row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        for entry in self._history:
            run_dir = entry.get('run_dir', '')
            if run_dir and os.path.isdir(run_dir):
                try: shutil.rmtree(run_dir)
                except Exception: pass
        self._history.clear()
        self._settings.setValue('run_history', '[]')
        self.accept()


# ── Main window ───────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('מתכנן הפערים')
        self.setMinimumWidth(660)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self._settings       = QSettings('TutorDiff', 'TutorDiff')
        self._auto_assigning = False
        self.setAcceptDrops(True)
        self._build_ui()
        self._restore_last_files()

    def _build_ui(self):
        root_widget = QWidget()
        root_widget.setStyleSheet(f'background: {BG};')
        self.setCentralWidget(root_widget)
        root = QVBoxLayout(root_widget)
        root.setContentsMargins(28, 22, 28, 28)
        root.setSpacing(14)

        # ── Gradient header card ─────────────────────────────────────────
        header = QWidget()
        header.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {PRIMARY}, stop:1 #2d5a8e);
                border-radius: 16px;
            }}
        """)
        header.setGraphicsEffect(_shadow(24, 7, 45))
        h_lay = QVBoxLayout(header)
        h_lay.setContentsMargins(24, 18, 24, 18)
        h_lay.setSpacing(4)

        title = QLabel('מתכנן הפערים')
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ft = QFont(); ft.setPointSize(22); ft.setBold(True)
        title.setFont(ft)
        title.setStyleSheet('color: #ffffff; background: transparent;')

        subtitle = QLabel('השווה קבצי Excel וצור דוח פערים מסודר')
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet('color: rgba(255,255,255,0.6); font-size: 12px; background: transparent;')

        h_lay.addWidget(title)
        h_lay.addWidget(subtitle)
        root.addWidget(header)

        # ── Drop zones ───────────────────────────────────────────────────
        zones_row = QHBoxLayout()
        zones_row.setSpacing(10)
        self.darush_zone = DropZone('דרוש תיקון', 'גרור לכאן את הקובץ לתיקון')
        self.yitzua_zone = DropZone('ייצוא שיטס',  'גרור לכאן את הקובץ התקין')
        self.darush_zone.file_changed.connect(self._on_darush_file)
        self.yitzua_zone.file_changed.connect(self._on_yitzua_file)
        self.darush_zone.cleared.connect(lambda: self._settings.setValue('darush_path', ''))
        self.yitzua_zone.cleared.connect(lambda: self._settings.setValue('yitzua_path', ''))

        swap_btn = QPushButton('⇄')
        swap_btn.setToolTip('החלף בין הקבצים')
        swap_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        swap_btn.setFixedSize(36, 36)
        swap_btn.setStyleSheet(f"""
            QPushButton {{
                background: #f1f5f9; border: 1.5px solid {BORDER};
                border-radius: 9px; font-size: 18px; color: {MUTED};
            }}
            QPushButton:hover   {{ background: #dbeafe; border-color: {ACCENT}; color: {ACCENT}; }}
            QPushButton:pressed {{ background: #bfdbfe; }}
        """)
        swap_btn.clicked.connect(self._swap_files)

        zones_row.addWidget(self.darush_zone)
        zones_row.addWidget(swap_btn, 0, Qt.AlignmentFlag.AlignVCenter)
        zones_row.addWidget(self.yitzua_zone)
        root.addLayout(zones_row)

        # ── Date range card ──────────────────────────────────────────────
        self.date_card = DateRangeCard()
        root.addWidget(self.date_card)

        # ── Progress bar ─────────────────────────────────────────────────
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(5)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                background: #dde3ea; border-radius: 2px; border: none;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {ACCENT}, stop:1 #93c5fd);
                border-radius: 2px;
            }}
        """)
        self.progress_bar.setVisible(False)
        root.addWidget(self.progress_bar)

        self.status_lbl = QLabel('')
        self.status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_lbl.setStyleSheet(f'color: {MUTED}; font-size: 12px; min-height: 16px;')
        root.addWidget(self.status_lbl)

        self.cancel_btn = QPushButton('✕   ביטול')
        self.cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.cancel_btn.setFixedHeight(30)
        self.cancel_btn.setStyleSheet("""
            QPushButton {
                background: #fee2e2; border: 1px solid #fca5a5;
                border-radius: 7px; padding: 3px 20px;
                font-size: 12px; color: #dc2626;
            }
            QPushButton:hover { background: #fca5a5; }
        """)
        self.cancel_btn.clicked.connect(self._cancel)
        self.cancel_btn.setVisible(False)
        root.addWidget(self.cancel_btn, 0, Qt.AlignmentFlag.AlignCenter)

        # ── Build button ─────────────────────────────────────────────────
        self.build_btn = QPushButton('✦   צור דוח')
        self.build_btn.setMinimumHeight(52)
        fb = QFont(); fb.setPointSize(16); fb.setBold(True)
        self.build_btn.setFont(fb)
        self.build_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.build_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {PRIMARY}, stop:1 {PRIMARY_H});
                color: white; border-radius: 13px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {PRIMARY_H}, stop:1 #3a6fa8);
            }}
            QPushButton:pressed {{ background: #0f2540; }}
            QPushButton:disabled {{ background: #94a3b8; color: #e2e8f0; }}
        """)
        self.build_btn.setGraphicsEffect(_shadow(12, 4, 35))
        self.build_btn.setToolTip('⌘↩')
        self.build_btn.clicked.connect(self._run)
        QShortcut(QKeySequence('Ctrl+Return'), self).activated.connect(self._run)
        root.addWidget(self.build_btn)

        # ── Results card ─────────────────────────────────────────────────
        self.results_card = ResultsCard()
        root.addWidget(self.results_card)

        # ── History button ───────────────────────────────────────────────
        self.history_btn = QPushButton('📋   היסטוריית ריצות')
        self.history_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.history_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; border: 1px solid {BORDER};
                border-radius: 8px; padding: 6px 18px;
                font-size: 12px; color: {MUTED};
            }}
            QPushButton:hover {{ background: white; color: {TEXT}; border-color: #94a3b8; }}
        """)
        self.history_btn.clicked.connect(self._show_history)
        root.addWidget(self.history_btn, 0, Qt.AlignmentFlag.AlignCenter)

    # ── helpers ───────────────────────────────────────────────────────────────

    def _restore_last_files(self):
        self._auto_assigning = True   # don't re-detect files from previous session
        try:
            for key, zone in [('darush_path', self.darush_zone),
                               ('yitzua_path', self.yitzua_zone)]:
                path = self._settings.value(key, '')
                if path and os.path.exists(path):
                    zone.set_file(path)
        finally:
            self._auto_assigning = False

    def _swap_files(self):
        d_path = self.darush_zone.file_path
        y_path = self.yitzua_zone.file_path
        self._auto_assigning = True   # user explicitly swapped — don't re-detect
        try:
            if y_path:
                self.darush_zone.set_file(y_path)
            else:
                self.darush_zone._clear()
            if d_path:
                self.yitzua_zone.set_file(d_path)
            else:
                self.yitzua_zone._clear()
        finally:
            self._auto_assigning = False

    # ── File-type auto-detection ──────────────────────────────────────────────

    def _on_darush_file(self, path: str):
        self._settings.setValue('darush_path', path)
        if not self._auto_assigning:
            self._auto_assign(path, self.darush_zone)

    def _on_yitzua_file(self, path: str):
        self._settings.setValue('yitzua_path', path)
        if not self._auto_assigning:
            self._auto_assign(path, self.yitzua_zone)

    def _auto_assign(self, path: str, dropped_on: DropZone):
        detected = _detect_file_type(path)
        if detected is None:
            return
        correct = self.darush_zone if detected == 'darush' else self.yitzua_zone
        if dropped_on is correct:
            return
        self._auto_assigning = True
        try:
            dropped_on._clear()
            correct.set_file(path)
            self.status_lbl.setText('✓ זוהה אוטומטית — סודר לאזור הנכון')
            QTimer.singleShot(3000, lambda: self.status_lbl.setText(''))
        finally:
            self._auto_assigning = False

    # ── Run history ───────────────────────────────────────────────────────────

    def _load_run_history(self) -> list:
        import json
        try:
            return json.loads(self._settings.value('run_history', '[]'))
        except Exception:
            return []

    def _save_run_history(self, entry: dict):
        import json, shutil
        base = QStandardPaths.writableLocation(
            QStandardPaths.StandardLocation.AppDataLocation)
        run_dir = os.path.join(base, 'runs',
                               datetime.now().strftime('%Y%m%d_%H%M%S'))
        try:
            os.makedirs(run_dir, exist_ok=True)
            for priv_key, pub_key, dest_name in [
                ('_report_path', 'report_path', 'דוח_פערים.xlsx'),
                ('_darush_path', 'darush_path', 'דרוש_תיקון.xlsx'),
                ('_yitzua_path', 'yitzua_path', 'ייצוא_שיטס.xlsx'),
            ]:
                src = entry.pop(priv_key, None)
                if src and os.path.exists(src):
                    dest = os.path.join(run_dir, dest_name)
                    shutil.copy2(src, dest)
                    entry[pub_key] = dest
            entry['run_dir'] = run_dir
        except Exception:
            pass

        history = self._load_run_history()
        history.insert(0, entry)
        for old in history[20:]:
            old_dir = old.get('run_dir', '')
            if old_dir and os.path.isdir(old_dir):
                try: shutil.rmtree(old_dir)
                except Exception: pass
        self._settings.setValue('run_history', json.dumps(history[:20]))

    def _show_history(self):
        HistoryDialog(self._load_run_history(), self._settings, self).exec()

    # ── run ───────────────────────────────────────────────────────────────────

    def _run(self):
        darush = self.darush_zone.file_path
        yitzua = self.yitzua_zone.file_path

        if not darush:
            _alert(self, 'חסר קובץ', 'יש לבחור קובץ "דרוש תיקון".')
            return
        if not yitzua:
            _alert(self, 'חסר קובץ', 'יש לבחור קובץ "ייצוא שיטס".')
            return
        if not os.path.exists(darush):
            _alert(self, 'קובץ לא נמצא', f'הקובץ "דרוש תיקון" לא נמצא בנתיב:\n{darush}')
            self.darush_zone._clear()
            return
        if not os.path.exists(yitzua):
            _alert(self, 'קובץ לא נמצא', f'הקובץ "ייצוא שיטס" לא נמצא בנתיב:\n{yitzua}')
            self.yitzua_zone._clear()
            return
        if os.path.abspath(darush) == os.path.abspath(yitzua):
            _alert(self, 'קבצים זהים', 'שני הקבצים הם אותו קובץ.\nיש לבחור שני קבצים שונים.')
            return
        if self.date_card.is_active():
            if self.date_card.get_start() > self.date_card.get_end():
                _alert(self, 'טווח תאריכים שגוי', 'תאריך ההתחלה חייב להיות לפני תאריך הסיום.')
                return

        # Build to a temp file first — we learn the date range only after processing
        tmp_output = os.path.join(tempfile.mkdtemp(), 'report.xlsx')

        self.results_card.hide_results()
        self.build_btn.setEnabled(False)
        self.cancel_btn.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.status_lbl.setText('מעבד…')

        self._worker = Worker(
            self.darush_zone.file_path,
            self.yitzua_zone.file_path,
            self.date_card.is_active(),
            date_start=self.date_card.get_start(),
            date_end=self.date_card.get_end(),
            output=tmp_output,
        )
        self._worker.finished.connect(self._on_done)
        self._worker.error.connect(self._on_error)
        self._worker.status.connect(self.status_lbl.setText)
        self._worker.progress.connect(self.progress_bar.setValue)
        self._worker.start()

    @staticmethod
    def _smart_filename(drange: str) -> str:
        """'01.05.2026 — 31.05.2026'  →  'דוח_פערים_01.05-31.05.xlsx'"""
        parts = [p.strip() for p in drange.split('—')]
        if len(parts) == 2:
            d1 = '.'.join(parts[0].split('.')[:2])   # "01.05"
            d2 = '.'.join(parts[1].split('.')[:2])   # "31.05"
            return f'דוח_פערים_{d1}-{d2}.xlsx'
        return 'דוח_פערים.xlsx'

    def _on_done(self, tmp_path: str, math_count: int, eng_count: int,
                 mark_count: int, warnings_count: int, drange: str):
        self.build_btn.setEnabled(True)
        self.cancel_btn.setVisible(False)
        self.progress_bar.setVisible(False)
        self.status_lbl.setText('')

        # Ask where to save, with smart filename derived from the actual date range
        from PyQt6.QtWidgets import QFileDialog
        filename = self._smart_filename(drange)
        default  = os.path.join(os.path.expanduser('~'), 'Downloads', filename)
        output, _ = QFileDialog.getSaveFileName(self, 'שמור דוח', default, 'Excel (*.xlsx)')
        tmp_dir = os.path.dirname(tmp_path)
        if not output:
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return
        if not output.endswith('.xlsx'):
            output += '.xlsx'

        shutil.move(tmp_path, output)
        shutil.rmtree(tmp_dir, ignore_errors=True)
        self.results_card.show_results(output, math_count, eng_count, mark_count, warnings_count)
        self._mac_notify(math_count + eng_count)
        self._save_run_history({
            'timestamp':      datetime.now().strftime('%d.%m.%Y %H:%M'),
            'drange':         drange,
            'math_count':     math_count,
            'eng_count':      eng_count,
            'mark_count':     mark_count,
            'warnings_count': warnings_count,
            'darush_file':    Path(self.darush_zone.file_path or '').name,
            'yitzua_file':    Path(self.yitzua_zone.file_path or '').name,
            '_report_path':   output,
            '_darush_path':   self.darush_zone.file_path,
            '_yitzua_path':   self.yitzua_zone.file_path,
        })

    def _on_error(self, msg: str):
        self.build_btn.setEnabled(True)
        self.cancel_btn.setVisible(False)
        self.progress_bar.setVisible(False)
        self.status_lbl.setText('')
        _alert(self, 'שגיאה', msg, 'error')

    def _cancel(self):
        if hasattr(self, '_worker') and self._worker.isRunning():
            # Disconnect signals first — terminate() is async and could still fire them
            for sig in (self._worker.finished, self._worker.error,
                        self._worker.status, self._worker.progress):
                try: sig.disconnect()
                except Exception: pass
            self._worker.terminate()
            self._worker.wait(1500)
            try: shutil.rmtree(os.path.dirname(self._worker.output), ignore_errors=True)
            except Exception: pass
        self.build_btn.setEnabled(True)
        self.cancel_btn.setVisible(False)
        self.progress_bar.setVisible(False)
        self.status_lbl.setText('')

    # ── Window-wide drag-and-drop ─────────────────────────────────────────────

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0].toLocalFile().lower()
            if url.endswith(('.xlsx', '.xls')):
                event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if not urls:
            return
        path = urls[0].toLocalFile()
        if not path.lower().endswith(('.xlsx', '.xls')):
            return
        detected = _detect_file_type(path)
        if detected == 'darush':
            self.darush_zone.set_file(path)
        elif detected == 'yitzua':
            self.yitzua_zone.set_file(path)
        elif not self.darush_zone.file_path:
            self.darush_zone.set_file(path)
        elif not self.yitzua_zone.file_path:
            self.yitzua_zone.set_file(path)
        else:
            self.darush_zone.set_file(path)

    def _mac_notify(self, total: int):
        try:
            subprocess.run([
                'osascript', '-e',
                f'display notification "נמצאו {total} פערים — הדוח מוכן" '
                f'with title "מתכנן הפערים" sound name "Glass"'
            ], check=False)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
    app.setStyle('Fusion')
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
